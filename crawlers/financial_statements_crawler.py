"""
财务报表爬虫 - 获取完整的三张财务报表数据
支持利润表、资产负债表、现金流量表
"""
import json
import time
from datetime import datetime
from typing import List, Dict, Any, Optional
from crawlers.base_crawler import BaseCrawler
from engine.logger import get_logger
from engine.database import DataRepository
from engine.xueqiu_auth import XueqiuAuth

logger = get_logger(__name__)


class FinancialStatementsCrawler(BaseCrawler):
    """财务报表爬虫"""
    
    def __init__(self, data_repo: DataRepository = None):
        self.data_repo = data_repo or DataRepository()
        self.auth = XueqiuAuth()
        self.session = self.auth.get_session()
        
        # 财务报表API端点 - 借鉴recycling/finstat.py
        self.statements_urls = {
            'income': 'https://stock.xueqiu.com/v5/stock/finance/cn/income.json',      # 利润表
            'balance': 'https://stock.xueqiu.com/v5/stock/finance/cn/balance.json',    # 资产负债表
            'cash': 'https://stock.xueqiu.com/v5/stock/finance/cn/cash_flow.json',   # 现金流量表
            'indicator': 'https://stock.xueqiu.com/v5/stock/finance/cn/indicator.json', # 主要指标
            # 备用API端点
            'income_backup': 'https://xueqiu.com/services/v10/stock/finance/lrb.json',
            'balance_backup': 'https://xueqiu.com/services/v10/stock/finance/zcfz.json', 
            'cash_backup': 'https://xueqiu.com/services/v10/stock/finance/xjll.json',
        }
        
        # 请求头 - 借鉴recycling/finstat.py
        self.headers = {
            'Host': 'stock.xueqiu.com',
            'Accept': 'application/json',
            'User-Agent': 'Xueqiu iPhone 11.8',
            'Accept-Language': 'zh-Hans-CN;q=1, ja-JP;q=0.9',
            'Accept-Encoding': 'br, gzip, deflate',
            'Connection': 'keep-alive'
        }
        
        logger.info("财务报表爬虫初始化完成")
    
    @staticmethod
    def safe_number(value):
        """安全数值转换 - 借鉴recycling/finstat.py"""
        if value is None or value == '':
            return 0
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0
    
    @staticmethod
    def format_percent(value, decimal_places=5):
        """格式化百分比 - 借鉴recycling/finstat.py"""
        if value is None or value == '':
            return "-"
        try:
            return round(float(value), decimal_places)
        except (ValueError, TypeError):
            return "-"
    
    @staticmethod
    def format_percent_100(value):
        """格式化100%百分比 - 借鉴recycling/finstat.py"""
        if value is None or value == '':
            return "-"
        try:
            return f"{round(float(value) * 100, 2)}%"
        except (ValueError, TypeError):
            return "-"
    
    @staticmethod
    def format_round(value, decimal_places=3):
        """格式化小数 - 借鉴recycling/finstat.py"""
        if value is None or value == '':
            return "-"
        try:
            return round(float(value), decimal_places)
        except (ValueError, TypeError):
            return "-"
    
    @staticmethod
    def format_yi(value):
        """转换为亿元单位 - 借鉴recycling/finstat.py"""
        if value is None or value == '':
            return "-"
        try:
            return round(float(value) / 10**8, 2)
        except (ValueError, TypeError):
            return "-"
    
    def _try_api_endpoints(self, statement_type: str, symbol: str, max_retries: int = 3) -> Optional[List[Dict[str, Any]]]:
        """
        尝试多个API端点获取财务报表数据
        
        Args:
            statement_type: 报表类型 ('income', 'balance', 'cash')
            symbol: 股票代码
            max_retries: 最大重试次数
            
        Returns:
            List[Dict]: 财务报表数据列表
        """
        # 获取该类型报表的所有可能端点
        primary_url = self.statements_urls.get(statement_type)
        backup_url = self.statements_urls.get(f'{statement_type}_backup')
        
        urls_to_try = []
        if primary_url:
            urls_to_try.append(primary_url)
        if backup_url:
            urls_to_try.append(backup_url)
        
        if not urls_to_try:
            logger.error(f"没有找到{statement_type}类型报表的可用API端点")
            return None
        
        timestamp = int(time.time() * 1000)
        
        for url in urls_to_try:
            for attempt in range(max_retries):
                try:
                    # 构建请求URL - 借鉴recycling/finstat.py使用timestamp
                    if 'cn/' in url:
                        # 使用v5/cn/ API格式
                        full_url = f"{url}?symbol={symbol}&type=all&is_detail=true&count=20&timestamp={timestamp}"
                    else:
                        # 使用services/v10/ API格式
                        full_url = f"{url}?symbol={symbol}&page=1&size=20&_={timestamp}"
                    
                    # 调整Host头
                    headers = self.headers.copy()
                    if 'stock.xueqiu.com' in url:
                        headers['Host'] = 'stock.xueqiu.com'
                    
                    logger.debug(f"尝试API: {url}")
                    response = self.session.get(full_url, headers=headers, timeout=10)
                    
                    if response.status_code == 200:
                        # 尝试解析JSON
                        text = response.text.replace('null', '0')
                        data = json.loads(text)
                        
                        # 检查数据结构
                        if 'list' in data and data['list']:
                            return data['list']
                        elif 'data' in data and data['data']:
                            if isinstance(data['data'], list):
                                return data['data']
                            elif isinstance(data['data'], dict) and 'list' in data['data']:
                                return data['data']['list']
                        else:
                            logger.warning(f"API返回数据结构异常: {list(data.keys())}")
                            
                    elif response.status_code == 404:
                        logger.debug(f"API端点不存在: {url}")
                        break  # 404说明端点不存在，尝试下一个URL
                    else:
                        logger.warning(f"API请求失败，状态码: {response.status_code}，尝试次数: {attempt + 1}")
                        
                except Exception as e:
                    logger.error(f"API请求异常: {e}，尝试次数: {attempt + 1}")
                
                # 重试前等待
                if attempt < max_retries - 1:
                    time.sleep(0.5)
                    timestamp = int(time.time() * 1000)  # 更新时间戳
        
        return None
    
    def get_financial_statement(self, symbol: str, statement_type: str) -> Optional[List[Dict[str, Any]]]:
        """
        获取指定类型的财务报表数据
        
        Args:
            symbol: 股票代码
            statement_type: 报表类型 ('income', 'balance', 'cash')
            
        Returns:
            List[Dict]: 财务报表数据列表
        """
        if statement_type not in ['income', 'balance', 'cash']:
            logger.error(f"不支持的报表类型: {statement_type}")
            return None
        
        logger.info(f"获取{symbol}的{statement_type}报表数据...")
        
        statement_data = self._try_api_endpoints(statement_type, symbol)
        
        if statement_data:
            # 处理数据格式
            processed_data = []
            for item in statement_data:
                # 用股票代码替换雪球内部编号
                item['symbol'] = symbol
                item['statement_type'] = statement_type
                item['crawl_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                item['timestamp'] = int(time.time())
                
                # 确保所有字段都有值，空值设为0
                for key, value in item.items():
                    if value is None or value == '':
                        item[key] = 0
                
                processed_data.append(item)
            
            logger.info(f"成功获取{symbol}的{statement_type}报表数据，共{len(processed_data)}条记录")
            return processed_data
        else:
            logger.warning(f"未获取到{symbol}的{statement_type}报表数据")
            return None
    
    def get_all_statements(self, symbol: str) -> Dict[str, List[Dict[str, Any]]]:
        """
        获取股票的所有三张财务报表
        
        Args:
            symbol: 股票代码
            
        Returns:
            Dict: 包含三张报表数据的字典
        """
        statements = {}
        statement_types = ['income', 'balance', 'cash']
        statement_names = ['利润表', '资产负债表', '现金流量表']
        
        for stmt_type, stmt_name in zip(statement_types, statement_names):
            logger.info(f"正在获取{symbol}的{stmt_name}...")
            
            try:
                data = self.get_financial_statement(symbol, stmt_type)
                if data:
                    statements[stmt_type] = data
                else:
                    logger.warning(f"{symbol}的{stmt_name}获取失败")
                    
            except Exception as e:
                logger.error(f"获取{symbol}的{stmt_name}异常: {e}")
            
            # 降低请求频率
            time.sleep(0.3)
        
        logger.info(f"{symbol}财务报表获取完成，成功获取{len(statements)}张报表")
        return statements
    
    def save_financial_statement(self, symbol: str, statement_type: str, statement_data: List[Dict[str, Any]]) -> bool:
        """
        保存财务报表数据
        
        Args:
            symbol: 股票代码
            statement_type: 报表类型
            statement_data: 报表数据列表
            
        Returns:
            bool: 是否成功
        """
        try:
            success_count = 0
            error_count = 0
            
            # CSV模式批量保存
            if hasattr(self.data_repo, 'csv_storage') and self.data_repo.csv_storage:
                try:
                    success = self.data_repo.csv_storage.save_financial_statement(
                        symbol, statement_type, statement_data
                    )
                    if success:
                        success_count += len(statement_data)
                        for data in statement_data:
                            report_date = data.get('reportdate', data.get('report_date', ''))
                            logger.info(f'{symbol}*{report_date}*{statement_type}报表 保存成功')
                    else:
                        error_count += len(statement_data)
                except Exception as e:
                    error_count += len(statement_data)
                    logger.error(f'批量保存{symbol}的{statement_type}报表失败: {e}')
            else:
                # 数据库模式逐条保存
                for data in statement_data:
                    try:
                        success = self.data_repo.save_financial_statement_data(symbol, statement_type, data)
                        if success:
                            success_count += 1
                            report_date = data.get('reportdate', data.get('report_date', ''))
                            logger.info(f'{symbol}*{report_date}*{statement_type}报表 保存成功')
                        else:
                            error_count += 1
                    except Exception as e:
                        error_count += 1
                        logger.error(f'保存{symbol}的{statement_type}报表数据失败: {e}')
            
            logger.info(f"{symbol}的{statement_type}报表数据处理完成，成功: {success_count}, 失败: {error_count}")
            return success_count > 0
            
        except Exception as e:
            logger.error(f"保存{symbol}的{statement_type}报表异常: {e}")
            return False
    
    def save_all_statements(self, symbol: str, statements: Dict[str, List[Dict[str, Any]]]) -> bool:
        """
        保存所有财务报表数据
        
        Args:
            symbol: 股票代码
            statements: 包含三张报表数据的字典
            
        Returns:
            bool: 是否成功
        """
        total_success = True
        
        for statement_type, data in statements.items():
            if data:
                success = self.save_financial_statement(symbol, statement_type, data)
                total_success = total_success and success
        
        return total_success
    
    def crawl_single_stock_statements(self, symbol: str) -> bool:
        """
        爬取单只股票的完整财务报表
        
        Args:
            symbol: 股票代码
            
        Returns:
            bool: 是否成功
        """
        logger.info(f"开始爬取{symbol}的完整财务报表...")
        
        try:
            # 获取所有报表
            statements = self.get_all_statements(symbol)
            
            if statements:
                # 保存所有报表
                return self.save_all_statements(symbol, statements)
            else:
                logger.warning(f"未获取到{symbol}的任何财务报表数据")
                return False
                
        except Exception as e:
            logger.error(f"爬取{symbol}财务报表失败: {e}")
            return False
    
    def get_unprocessed_stocks(self) -> List[str]:
        """
        获取未处理财务报表的股票列表
        
        Returns:
            List[str]: 股票代码列表
        """
        try:
            if hasattr(self.data_repo, 'csv_storage') and self.data_repo.csv_storage:
                # CSV模式
                all_stocks = self.data_repo.csv_storage.get_latest_stock_info()
                processed_logs = self.data_repo.csv_storage.get_financial_statement_logs()
                
                # 构建已处理的股票-报表类型集合
                processed_set = set()
                for log in processed_logs:
                    symbol = log.get('symbol', '')
                    stmt_type = log.get('statement_type', '')
                    if symbol and stmt_type:
                        processed_set.add(f"{symbol}_{stmt_type}")
                
                # 返回未处理完整三表的股票
                unprocessed_symbols = []
                for stock in all_stocks:
                    symbol = stock.get('symbol', '')
                    if symbol:
                        # 检查是否三张表都已处理
                        has_all_statements = all(
                            f"{symbol}_{stmt_type}" in processed_set 
                            for stmt_type in ['income', 'balance', 'cash']
                        )
                        if not has_all_statements:
                            unprocessed_symbols.append(symbol)
                
                return unprocessed_symbols[:30]  # 限制数量
                
            else:
                # 数据库模式
                query = '''
                SELECT DISTINCT a.symbol 
                FROM stocks a 
                WHERE a.symbol NOT IN (
                    SELECT DISTINCT symbol FROM financial_statements_log 
                    WHERE statement_type IN ('income', 'balance', 'cash')
                    GROUP BY symbol 
                    HAVING COUNT(DISTINCT statement_type) = 3
                )
                ORDER BY a.symbol ASC 
                LIMIT 30
                '''
                result = self.data_repo.database.execute_query(query)
                return [row.get('symbol', '') for row in result if row.get('symbol')]
                
        except Exception as e:
            logger.error(f"获取未处理股票列表失败: {e}")
            return []
    
    def crawl_financial_statements(self):
        """爬取财务报表主函数"""
        logger.info("开始爬取财务报表数据...")
        
        # 获取未处理的股票列表
        stock_symbols = self.get_unprocessed_stocks()
        total = len(stock_symbols)
        
        if total == 0:
            logger.info("没有需要处理财务报表的股票")
            return
        
        logger.info(f"找到{total}支股票需要处理财务报表")
        
        success_count = 0
        error_count = 0
        
        for i, symbol in enumerate(stock_symbols, 1):
            logger.info(f'第{i}支股票，{symbol}')
            
            try:
                if self.crawl_single_stock_statements(symbol):
                    success_count += 1
                else:
                    error_count += 1
                    
            except Exception as e:
                error_count += 1
                logger.error(f'处理{symbol}财务报表异常: {e}')
            
            # 降低爬取速度
            time.sleep(0.5)
            
            logger.info('-' * 50)
        
        logger.info(f"财务报表爬取完成，总计: {total}, 成功: {success_count}, 失败: {error_count}")
    
    def calculate_financial_ratios(self, balance_data: Dict, income_data: Dict, cash_flow_data: Dict = None) -> Dict:
        """
        计算财务分析指标 - 借鉴recycling/finstat.py
        
        Args:
            balance_data: 资产负债表数据
            income_data: 利润表数据
            cash_flow_data: 现金流量表数据
            
        Returns:
            Dict: 计算得到的财务指标
        """
        ratios = {}
        
        try:
            # 资产负债表数据提取
            total_assets = self.safe_number(balance_data.get('total_assets', [0])[0])
            total_current_assets = self.safe_number(balance_data.get('total_current_assets', [0])[0])
            total_current_liab = self.safe_number(balance_data.get('total_current_liab', [0])[0])
            total_holders_equity = self.safe_number(balance_data.get('total_holders_equity', [0])[0])
            
            # 金融资产计算
            tradable_fnncl_assets = self.safe_number(balance_data.get('tradable_fnncl_assets', [0])[0])
            interest_receivable = self.safe_number(balance_data.get('interest_receivable', [0])[0])
            dividend_receivable = self.safe_number(balance_data.get('dividend_receivable', [0])[0])
            to_sale_asset = self.safe_number(balance_data.get('to_sale_asset', [0])[0])
            nca_due_within_one_year = self.safe_number(balance_data.get('nca_due_within_one_year', [0])[0])
            salable_financial_assets = self.safe_number(balance_data.get('salable_financial_assets', [0])[0])
            held_to_maturity_invest = self.safe_number(balance_data.get('held_to_maturity_invest', [0])[0])
            other_eq_ins_invest = self.safe_number(balance_data.get('other_eq_ins_invest', [0])[0])
            other_illiquid_fnncl_assets = self.safe_number(balance_data.get('other_illiquid_fnncl_assets', [0])[0])
            invest_property = self.safe_number(balance_data.get('invest_property', [0])[0])
            
            # 长期股权投资
            lt_equity_invest = self.safe_number(balance_data.get('lt_equity_invest', [0])[0])
            
            # 固定资产相关
            fixed_asset_sum = self.safe_number(balance_data.get('fixed_asset_sum', [0])[0])
            construction_in_process_sum = self.safe_number(balance_data.get('construction_in_process_sum', [0])[0])
            intangible_assets = self.safe_number(balance_data.get('intangible_assets', [0])[0])
            goodwill = self.safe_number(balance_data.get('goodwill', [0])[0])
            
            # 利润表数据提取
            total_revenue = self.safe_number(income_data.get('total_revenue', [0])[0])
            operating_cost = self.safe_number(income_data.get('operating_cost', [0])[0])
            operating_costs = self.safe_number(income_data.get('operating_costs', [0])[0])
            net_profit = self.safe_number(income_data.get('net_profit', [0])[0])
            income_tax_expenses = self.safe_number(income_data.get('income_tax_expenses', [0])[0])
            financing_expenses = self.safe_number(income_data.get('financing_expenses', [0])[0])
            profit_total_amt = self.safe_number(income_data.get('profit_total_amt', [0])[0])
            
            # 计算金融资产
            finance_asset = (tradable_fnncl_assets + interest_receivable + dividend_receivable + 
                           to_sale_asset + nca_due_within_one_year + salable_financial_assets + 
                           held_to_maturity_invest + other_eq_ins_invest + other_illiquid_fnncl_assets + 
                           invest_property)
            
            # 计算经营资产
            business_asset = total_assets - finance_asset - lt_equity_invest
            
            # 计算长期经营资产
            long_business_asset = (fixed_asset_sum + construction_in_process_sum + 
                                 intangible_assets + goodwill)
            
            # 计算各项比率
            if total_assets > 0:
                ratios['finance_asset_percent'] = finance_asset / total_assets
                ratios['business_asset_percent'] = business_asset / total_assets
                ratios['lt_equity_invest_percent'] = lt_equity_invest / total_assets
                ratios['long_business_asset_percent'] = long_business_asset / total_assets
            
            if total_holders_equity > 0:
                ratios['z_business_asset_percent'] = (total_current_assets - total_current_liab) / total_holders_equity
                ratios['finance_turnover'] = total_assets / total_holders_equity  # 财务杠杆倍数
            
            if (total_revenue - operating_costs) > 0:
                ratios['business_lever'] = (total_revenue - operating_cost) / (total_revenue - operating_costs)
            
            if total_revenue > 0:
                ratios['per_benefit_percent'] = (net_profit + income_tax_expenses + financing_expenses) / total_revenue
            
            # 计算周转率（需要根据报告期调整）
            report_name = income_data.get('report_name', '')
            days = 365
            if '一季报' in report_name:
                days = 90
            elif '中报' in report_name:
                days = 181
            elif '三季报' in report_name:
                days = 273
            
            times = 365 / days
            
            if long_business_asset > 0:
                ratios['long_asset_turnover_percent'] = (total_revenue / long_business_asset) * times
            
            if (total_current_assets - total_current_liab) > 0:
                ratios['business_turnover_percent'] = (total_revenue / (total_current_assets - total_current_liab)) * times
            
            # 财务成本效应比率
            if (net_profit + income_tax_expenses + financing_expenses) > 0:
                ratios['financial_cost_percent'] = (net_profit + income_tax_expenses) / (net_profit + income_tax_expenses + financing_expenses)
            
            # 企业所得税效应比率
            if (net_profit + income_tax_expenses) > 0:
                ratios['tax_cost_percent'] = net_profit / (net_profit + income_tax_expenses)
            
            # 扣非净利率占比
            if profit_total_amt > 0:
                net_profit_after_nrgal_atsolc = self.safe_number(income_data.get('net_profit_after_nrgal_atsolc', [0])[0])
                ratios['real_benefit_percent'] = net_profit_after_nrgal_atsolc / profit_total_amt
            
            # 现金流分析（如果有现金流量表数据）
            if cash_flow_data:
                ncf_from_oa = self.safe_number(cash_flow_data.get('ncf_from_oa', [0])[0])
                cash_received_of_sales_service = self.safe_number(cash_flow_data.get('cash_received_of_sales_service', [0])[0])
                cash_paid_for_assets = self.safe_number(cash_flow_data.get('cash_paid_for_assets', [0])[0])
                net_cash_of_disposal_assets = self.safe_number(cash_flow_data.get('net_cash_of_disposal_assets', [0])[0])
                
                if profit_total_amt > 0:
                    ratios['benefit_percentage'] = ncf_from_oa / profit_total_amt  # 净利润检验含金量
                
                if total_revenue > 0:
                    ratios['op_percentage'] = cash_received_of_sales_service / total_revenue  # 营业收入含金量检验
                
                # 长期经营资产净投资额
                net_investment = cash_paid_for_assets - net_cash_of_disposal_assets
                ratios['net_investment'] = net_investment
                
                # 长期经营资产扩张性资本支出
                ratios['out_lay'] = net_investment  # 简化版本，不考虑折旧
                
                if long_business_asset > 0:
                    ratios['out_lay_percent'] = ratios['out_lay'] / long_business_asset
                
                # 并购活动净合并额
                net_cash_amt_from_branch = self.safe_number(cash_flow_data.get('net_cash_amt_from_branch', [0])[0])
                net_cash_of_disposal_branch = self.safe_number(cash_flow_data.get('net_cash_of_disposal_branch', [0])[0])
                ratios['net_merger'] = net_cash_amt_from_branch - net_cash_of_disposal_branch
                
                # 现金自给率
                if (net_investment + ratios['net_merger']) != 0:
                    ratios['cash_satify_percent'] = ncf_from_oa / (net_investment + ratios['net_merger'])
            
            # 存储原始数据
            ratios.update({
                'finance_asset': finance_asset,
                'business_asset': business_asset,
                'lt_equity_invest': lt_equity_invest,
                'goodwill': goodwill,
                'long_business_asset': long_business_asset,
                'total_current_assets': total_current_assets,
                'total_current_liab': total_current_liab,
                'total_revenue': total_revenue,
                'net_profit': net_profit,
                'profit_total_amt': profit_total_amt
            })
            
        except Exception as e:
            logger.error(f"计算财务指标失败: {e}")
        
        return ratios
    
    def generate_excel_report(self, symbol: str, stock_name: str, financial_ratios: Dict, output_path: str = None) -> bool:
        """
        生成Excel财务分析报表 - 借鉴recycling/finstat.py
        
        Args:
            symbol: 股票代码
            stock_name: 股票名称
            financial_ratios: 财务指标数据
            output_path: 输出文件路径
            
        Returns:
            bool: 是否成功
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            
            # 如果没有指定输出路径，使用默认路径
            if not output_path:
                safe_name = stock_name.replace('*', '').replace('/', '_')
                output_path = f"./{safe_name}-财务分析报表.xlsx"
            
            # 尝试加载模板，如果不存在则创建新工作簿
            try:
                wb = openpyxl.load_workbook('财务数据模板.xlsx')
                sht = wb['公司分析']
            except:
                wb = openpyxl.Workbook()
                sht = wb.active
                sht.title = '公司分析'
                # 创建表头
                headers = ['指标', '报告期1', '报告期2', '报告期3', '报告期4', '报告期5']
                for i, header in enumerate(headers, 1):
                    sht.cell(1, i, header)
            
            # 设置字体样式
            green_font = Font(size=16, color="00a600", bold=True)
            red_font = Font(size=16, color="ff0000", bold=True)
            del_font = Font(strike=True)
            
            # 填充数据
            col = 3  # 从第3列开始填充数据
            
            for report_date, ratios in financial_ratios.items():
                # 基本信息
                sht.cell(1, col, ratios.get('report_name', ''))
                
                # 资产结构指标
                sht.cell(2, col, self.format_percent(ratios.get('z_business_asset_percent', 0)))
                sht.cell(3, col, self.format_percent(ratios.get('long_business_asset_percent', 0)))
                sht.cell(4, col, self.format_round(ratios.get('fixed_asset_turnover_ratio', 0)))
                
                # 资产类型判断
                if ratios.get('long_business_asset_percent', 0) > 0.5:
                    sht.cell(5, col, '重资产')
                else:
                    sht.cell(5, col, '轻资产')
                    sht.cell(4, col).font = del_font
                
                # 资产构成
                sht.cell(6, col, f"{self.format_yi(ratios.get('finance_asset', 0))} ({self.format_percent_100(ratios.get('finance_asset_percent', 0))})")
                sht.cell(7, col, f"{self.format_yi(ratios.get('lt_equity_invest', 0))} ({self.format_percent_100(ratios.get('lt_equity_invest_percent', 0))})")
                sht.cell(8, col, f"{self.format_yi(ratios.get('business_asset', 0))} ({self.format_percent_100(ratios.get('business_asset_percent', 0))})")
                
                # 其他指标
                sht.cell(10, col, self.format_yi(ratios.get('goodwill', 0)))
                sht.cell(11, col, self.format_round(ratios.get('business_lever', 0)))
                sht.cell(12, col, self.format_round(ratios.get('cash_cycle', 0)))
                sht.cell(13, col, self.format_percent(ratios.get('per_benefit_percent', 0)))
                sht.cell(14, col, self.format_percent(ratios.get('long_asset_turnover_percent', 0)))
                sht.cell(15, col, self.format_percent(ratios.get('business_turnover_percent', 0)))
                sht.cell(16, col, self.format_percent(ratios.get('financial_cost_percent', 0)))
                sht.cell(17, col, self.format_round(ratios.get('finance_turnover', 0)))
                sht.cell(18, col, self.format_percent(ratios.get('tax_cost_percent', 0)))
                sht.cell(19, col, self.format_round(ratios.get('benefit_percentage', 0)))
                sht.cell(20, col, self.format_round(ratios.get('op_percentage', 0)))
                sht.cell(21, col, self.format_percent(ratios.get('real_benefit_percent', 0)))
                
                # 投资分析
                sht.cell(22, col, self.format_yi(ratios.get('net_investment', 0)))
                sht.cell(23, col, self.format_yi(ratios.get('out_lay', 0)))
                sht.cell(24, col, self.format_percent(ratios.get('out_lay_percent', 0)))
                sht.cell(25, col, self.format_yi(ratios.get('net_merger', 0)))
                sht.cell(26, col, self.format_percent(ratios.get('cash_satify_percent', 0)))
                sht.cell(29, col, self.format_yi(ratios.get('long_business_asset', 0)))
                
                # 设置单元格格式
                percentage_cells = [2, 3, 13, 14, 15, 16, 18, 21, 24, 26]
                for row in percentage_cells:
                    sht.cell(row, col).number_format = '0.00%'
                
                # 设置对齐方式
                sht.cell(1, col).alignment = Alignment(horizontal='center', vertical='center')
                sht.cell(5, col).alignment = Alignment(horizontal='right', vertical='center')
                for row in [6, 7, 8]:
                    sht.cell(row, col).alignment = Alignment(horizontal='right', vertical='center')
                
                col += 1
            
            # 保存文件
            wb.save(output_path)
            logger.info(f"Excel报表生成成功: {output_path}")
            return True
            
        except ImportError:
            logger.warning("未安装openpyxl，无法生成Excel报表")
            return False
        except Exception as e:
            logger.error(f"生成Excel报表失败: {e}")
            return False
    
    def get_comprehensive_financial_analysis(self, symbol: str) -> Dict:
        """
        获取综合财务分析数据 - 整合三张报表数据并计算指标
        
        Args:
            symbol: 股票代码
            
        Returns:
            Dict: 综合财务分析数据
        """
        try:
            logger.info(f"开始获取{symbol}的综合财务分析数据...")
            
            # 获取三张报表数据
            statements = self.get_all_statements(symbol)
            
            if not statements:
                logger.warning(f"未获取到{symbol}的财务报表数据")
                return {}
            
            # 获取股票基本信息
            stock_info = self._get_stock_quote(symbol)
            
            # 整合数据并计算指标
            comprehensive_data = {
                'stock_info': stock_info,
                'financial_ratios': {}
            }
            
            # 按报告期整合数据
            report_periods = set()
            
            # 收集所有报告期
            for stmt_type, data_list in statements.items():
                for item in data_list:
                    report_date = item.get('report_date', item.get('reportdate', ''))
                    if report_date:
                        report_periods.add(report_date)
            
            # 为每个报告期计算财务指标
            for report_date in sorted(report_periods, reverse=True):
                # 查找对应报告期的各类报表数据
                balance_data = {}
                income_data = {}
                cash_flow_data = {}
                
                # 查找资产负债表数据
                if 'balance' in statements:
                    for item in statements['balance']:
                        if item.get('report_date', item.get('reportdate', '')) == report_date:
                            balance_data = item
                            break
                
                # 查找利润表数据
                if 'income' in statements:
                    for item in statements['income']:
                        if item.get('report_date', item.get('reportdate', '')) == report_date:
                            income_data = item
                            break
                
                # 查找现金流量表数据
                if 'cash_flow' in statements:
                    for item in statements['cash_flow']:
                        if item.get('report_date', item.get('reportdate', '')) == report_date:
                            cash_flow_data = item
                            break
                
                # 计算财务指标
                if balance_data and income_data:
                    ratios = self.calculate_financial_ratios(balance_data, income_data, cash_flow_data)
                    ratios['report_date'] = report_date
                    ratios['report_name'] = income_data.get('report_name', balance_data.get('report_name', ''))
                    comprehensive_data['financial_ratios'][report_date] = ratios
            
            logger.info(f"{symbol}综合财务分析数据获取完成，共{len(comprehensive_data['financial_ratios'])}个报告期")
            return comprehensive_data
            
        except Exception as e:
            logger.error(f"获取{symbol}综合财务分析数据失败: {e}")
            return {}
    
    def _get_stock_quote(self, symbol: str) -> Dict:
        """
        获取股票基本信息 - 借鉴recycling/finstat.py
        
        Args:
            symbol: 股票代码
            
        Returns:
            Dict: 股票基本信息
        """
        try:
            timestamp = int(time.time() * 1000)
            url = f"https://stock.xueqiu.com/v5/stock/quote.json?extend=detail&symbol={symbol}&timestamp={timestamp}"
            
            response = self.session.get(url, headers=self.headers, timeout=10)
            
            if response.status_code == 200:
                data = json.loads(response.text)
                quote_data = data.get('data', {}).get('quote', {})
                
                return {
                    'symbol': symbol,
                    'name': quote_data.get('name', ''),
                    'current': quote_data.get('current', 0),
                    'percent': quote_data.get('percent', 0)
                }
            else:
                logger.warning(f"获取{symbol}股票信息失败，状态码: {response.status_code}")
                return {'symbol': symbol, 'name': '', 'current': 0, 'percent': 0}
                
        except Exception as e:
            logger.error(f"获取{symbol}股票信息异常: {e}")
            return {'symbol': symbol, 'name': '', 'current': 0, 'percent': 0}
    
    def generate_financial_analysis_report(self, symbol: str, generate_excel: bool = True) -> bool:
        """
        生成完整的财务分析报告
        
        Args:
            symbol: 股票代码
            generate_excel: 是否生成Excel报告
            
        Returns:
            bool: 是否成功
        """
        try:
            logger.info(f"开始生成{symbol}的财务分析报告...")
            
            # 获取综合财务分析数据
            analysis_data = self.get_comprehensive_financial_analysis(symbol)
            
            if not analysis_data:
                logger.warning(f"无法获取{symbol}的财务分析数据")
                return False
            
            stock_info = analysis_data.get('stock_info', {})
            financial_ratios = analysis_data.get('financial_ratios', {})
            
            if not financial_ratios:
                logger.warning(f"{symbol}没有可用的财务指标数据")
                return False
            
            # 生成Excel报告
            if generate_excel:
                success = self.generate_excel_report(
                    symbol, 
                    stock_info.get('name', symbol), 
                    financial_ratios
                )
                if success:
                    logger.info(f"{symbol}的Excel财务分析报告生成成功")
                else:
                    logger.warning(f"{symbol}的Excel财务分析报告生成失败")
            
            # 保存分析数据到数据库/CSV
            try:
                self._save_financial_analysis(symbol, analysis_data)
            except Exception as e:
                logger.warning(f"保存{symbol}财务分析数据失败: {e}")
            
            logger.info(f"{symbol}财务分析报告生成完成")
            return True
            
        except Exception as e:
            logger.error(f"生成{symbol}财务分析报告失败: {e}")
            return False
    
    def _save_financial_analysis(self, symbol: str, analysis_data: Dict):
        """
        保存财务分析数据
        
        Args:
            symbol: 股票代码
            analysis_data: 分析数据
        """
        try:
            if hasattr(self.data_repo, 'csv_storage') and self.data_repo.csv_storage:
                # CSV模式
                self.data_repo.csv_storage.save_financial_analysis(symbol, analysis_data)
            else:
                # 数据库模式
                for report_date, ratios in analysis_data.get('financial_ratios', {}).items():
                    ratios['symbol'] = symbol
                    ratios['report_date'] = report_date
                    ratios['crawl_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    self.data_repo.save_financial_analysis_data(ratios)
                    
        except Exception as e:
            logger.error(f"保存财务分析数据失败: {e}")